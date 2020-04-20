import numpy as np
import openpyxl


def FancyPrint(A, B, selected):
    for row in range(len(B)):
        print("(", end='')
        for col in range(len(A[row])):
            print("\t{1:10.2f}{0}".format(" " if (selected is None or selected != (row, col)) else "*", A[row][col]),
                  end='')
        print("\t) * (\tX{0}) = (\t{1:10.2f})".format(row + 1, B[row]))


wb = openpyxl.load_workbook('D:\Gauss.xlsx')
sheet = wb.get_sheet_by_name('Лист1')

f = open('D:\gauss_slv.txt', 'w')
row_ex = [1, 6, 11, 16, 21]

for row in row_ex:
    if row == 21:
        myA = np.empty([6, 6], dtype=float)

        for i in range(0, 6):
            for j in range(0, 6):
                myA.itemset((i, j), float(sheet.cell(row=(row + i), column=(j + 1)).value))

        myB = np.empty(6, dtype=float)

        for i in range(0, 6):
            myB.itemset(i, float(sheet.cell(row=(row + i), column=7).value))

    else:
        myA = np.empty([4, 4], dtype=float)

        for i in range(0, 4):
            for j in range(0, 4):
                myA.itemset((i, j), float(sheet.cell(row=(row + i), column=(j + 1)).value))

        myB = np.empty(4, dtype=float)

        for i in range(0, 4):
            myB.itemset(i, float(sheet.cell(row=(row + i), column=5).value))

    print("Исходная система:")
    FancyPrint(myA, myB, None)
    slv = np.linalg.solve(myA, myB)

    print("Решаем:")
    print(slv)
    f.write(str(slv) + '\n')
    print("\n\n")

f.close()
wb.close()
